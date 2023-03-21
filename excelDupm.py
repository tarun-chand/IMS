from openpyxl import load_workbook
import psycopg2 
conn = psycopg2.connect(host="localhost",database='durg_ims',user='postgres',password='123')
cursor =conn.cursor()
path = '../Masterdata/NEWpromodel.xlsx'
wb_obj = load_workbook(path)
sheet_obj = wb_obj.active
max_row = len(sheet_obj['A'])
max_col =2
l = []
for r in range(1,max_row+1):
    col=[]
    for c in range(1,max_col+1):
        cell_obj = sheet_obj.cell(row=r,column=c)
        col.insert(c,cell_obj.value)
    l.append(tuple(col))
# print(l)
q = "insert into store_productmodelmaster(product_mod_name,product_com_id_id) values(%s,%s)"
# print(q)
cursor.executemany(q,l)
conn.commit()
conn.close()